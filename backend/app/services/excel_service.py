"""
Excel Export Service for Call Sheets
Generates professionally formatted Excel files from call sheet data
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    Border,
    Side,
    PatternFill,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any, Optional


# Color scheme
HEADER_BG = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
SUBHEADER_BG = PatternFill(start_color="2d2d2d", end_color="2d2d2d", fill_type="solid")
ALT_ROW_BG = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
ACCENT_BG = PatternFill(start_color="ffd700", end_color="ffd700", fill_type="solid")
SAFETY_BG = PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid")

# Font styles
TITLE_FONT = Font(name='Arial', size=16, bold=True, color="FFFFFF")
HEADER_FONT = Font(name='Arial', size=11, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name='Arial', size=10, bold=True)
BODY_FONT = Font(name='Arial', size=10)
LABEL_FONT = Font(name='Arial', size=9, color="666666")

# Border styles
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)


def format_time_for_excel(time_str: str) -> str:
    """Format time string for Excel display"""
    if not time_str:
        return ""
    return time_str


def format_date_for_excel(date_str: str) -> str:
    """Format date string for Excel display"""
    if not date_str:
        return ""
    try:
        date = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        return date.strftime('%A, %B %d, %Y')
    except:
        return date_str


async def generate_call_sheet_excel(
    call_sheet: Dict[str, Any],
    people: List[Dict[str, Any]],
    scenes: List[Dict[str, Any]],
    locations: List[Dict[str, Any]],
    project: Dict[str, Any]
) -> bytes:
    """
    Generate an Excel workbook from call sheet data

    Args:
        call_sheet: The main call sheet data
        people: List of people (cast and crew)
        scenes: List of scenes
        locations: List of locations
        project: Project information

    Returns:
        bytes: The Excel file as bytes
    """
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create sheets
    overview_sheet = wb.create_sheet("Overview", 0)
    cast_sheet = wb.create_sheet("Cast & Crew", 1)
    scenes_sheet = wb.create_sheet("Scenes", 2)
    locations_sheet = wb.create_sheet("Locations", 3)

    # Remove the default sheet
    wb.remove(default_sheet)

    # Populate sheets
    _build_overview_sheet(overview_sheet, call_sheet, project, len(people), len(scenes), len(locations))
    _build_cast_crew_sheet(cast_sheet, people)
    _build_scenes_sheet(scenes_sheet, scenes)
    _build_locations_sheet(locations_sheet, locations)

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def _build_overview_sheet(ws, call_sheet: Dict, project: Dict, people_count: int, scenes_count: int, locations_count: int):
    """Build the overview sheet with call sheet summary"""

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30

    row = 1

    # Title Row
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1, value=call_sheet.get('title', 'Call Sheet'))
    cell.font = TITLE_FONT
    cell.fill = HEADER_BG
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30
    row += 1

    # Project name
    if project.get('title'):
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws.cell(row=row, column=1, value=project.get('title'))
        cell.font = HEADER_FONT
        cell.fill = SUBHEADER_BG
        cell.alignment = Alignment(horizontal='center')
        row += 1

    # Blank row
    row += 1

    # Date and Day Number
    ws.cell(row=row, column=1, value="Shoot Date").font = LABEL_FONT
    ws.cell(row=row, column=2, value=format_date_for_excel(call_sheet.get('shoot_date', '')))
    ws.cell(row=row, column=3, value="Day Number").font = LABEL_FONT
    ws.cell(row=row, column=4, value=call_sheet.get('day_number', ''))
    row += 1

    # Blank row
    row += 1

    # Call Times Section
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1, value="CALL TIMES")
    cell.font = SUBHEADER_FONT
    cell.fill = ACCENT_BG
    row += 1

    call_times = [
        ("Crew Call", call_sheet.get('crew_call', '')),
        ("Talent Call", call_sheet.get('talent_call', '')),
        ("Breakfast", call_sheet.get('breakfast_time', '')),
        ("Lunch", call_sheet.get('lunch_time', '')),
        ("Est. Wrap", call_sheet.get('wrap_time', '')),
        ("Sunrise", call_sheet.get('sunrise', '')),
        ("Sunset", call_sheet.get('sunset', '')),
    ]

    col = 1
    for label, value in call_times:
        if value:
            ws.cell(row=row, column=col, value=label).font = LABEL_FONT
            ws.cell(row=row, column=col + 1, value=value).font = BODY_FONT
            col += 2
            if col > 4:
                col = 1
                row += 1

    if col != 1:
        row += 1

    # Blank row
    row += 1

    # Summary Section
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1, value="SUMMARY")
    cell.font = SUBHEADER_FONT
    cell.fill = ACCENT_BG
    row += 1

    ws.cell(row=row, column=1, value="Cast & Crew").font = LABEL_FONT
    ws.cell(row=row, column=2, value=f"{people_count} people").font = BODY_FONT
    ws.cell(row=row, column=3, value="Scenes").font = LABEL_FONT
    ws.cell(row=row, column=4, value=f"{scenes_count} scenes").font = BODY_FONT
    row += 1

    ws.cell(row=row, column=1, value="Locations").font = LABEL_FONT
    ws.cell(row=row, column=2, value=f"{locations_count} locations").font = BODY_FONT
    row += 1

    # Blank row
    row += 1

    # Weather (if available)
    if call_sheet.get('weather_forecast'):
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws.cell(row=row, column=1, value="WEATHER")
        cell.font = SUBHEADER_FONT
        cell.fill = ACCENT_BG
        row += 1

        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=1, value=call_sheet.get('weather_forecast', '')).font = BODY_FONT
        row += 2

    # Safety Notes
    if call_sheet.get('safety_notes'):
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws.cell(row=row, column=1, value="SAFETY NOTES")
        cell.font = SUBHEADER_FONT
        cell.fill = SAFETY_BG
        row += 1

        ws.merge_cells(f'A{row}:D{row}')
        cell = ws.cell(row=row, column=1, value=call_sheet.get('safety_notes', ''))
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True)
        row += 2

    # General Notes
    if call_sheet.get('notes'):
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws.cell(row=row, column=1, value="NOTES")
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_BG
        cell.font = HEADER_FONT
        row += 1

        ws.merge_cells(f'A{row}:D{row}')
        cell = ws.cell(row=row, column=1, value=call_sheet.get('notes', ''))
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True)


def _build_cast_crew_sheet(ws, people: List[Dict]):
    """Build the cast and crew sheet"""

    # Set column widths
    columns = [
        ('A', 'Name', 25),
        ('B', 'Role/Position', 25),
        ('C', 'Department', 15),
        ('D', 'Call Time', 12),
        ('E', 'Phone', 15),
        ('F', 'Email', 25),
        ('G', 'Notes', 30),
    ]

    for col_letter, _, width in columns:
        ws.column_dimensions[col_letter].width = width

    # Header row
    row = 1
    for idx, (_, header, _) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_BG
        cell.border = THIN_BORDER
    row += 1

    # Separate cast and crew
    cast = [p for p in people if p.get('is_cast')]
    crew = [p for p in people if not p.get('is_cast')]

    # Cast section
    if cast:
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row=row, column=1, value="CAST")
        cell.font = SUBHEADER_FONT
        cell.fill = ACCENT_BG
        row += 1

        for person in cast:
            ws.cell(row=row, column=1, value=person.get('name', '')).font = BODY_FONT
            ws.cell(row=row, column=2, value=person.get('role_or_position', '')).font = BODY_FONT
            ws.cell(row=row, column=3, value=person.get('department', '')).font = BODY_FONT
            ws.cell(row=row, column=4, value=person.get('call_time', '')).font = BODY_FONT
            ws.cell(row=row, column=5, value=person.get('phone', '')).font = BODY_FONT
            ws.cell(row=row, column=6, value=person.get('email', '')).font = BODY_FONT
            ws.cell(row=row, column=7, value=person.get('notes', '')).font = BODY_FONT

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        row += 1  # Blank row

    # Crew section
    if crew:
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row=row, column=1, value="CREW")
        cell.font = SUBHEADER_FONT
        cell.fill = ACCENT_BG
        row += 1

        # Group by department
        departments = {}
        for person in crew:
            dept = person.get('department', 'Other')
            if dept not in departments:
                departments[dept] = []
            departments[dept].append(person)

        for dept_name, dept_members in sorted(departments.items()):
            # Department header
            ws.merge_cells(f'A{row}:G{row}')
            cell = ws.cell(row=row, column=1, value=dept_name)
            cell.font = Font(name='Arial', size=10, bold=True, italic=True)
            cell.fill = ALT_ROW_BG
            row += 1

            for person in dept_members:
                ws.cell(row=row, column=1, value=person.get('name', '')).font = BODY_FONT
                ws.cell(row=row, column=2, value=person.get('role_or_position', '')).font = BODY_FONT
                ws.cell(row=row, column=3, value=person.get('department', '')).font = BODY_FONT
                ws.cell(row=row, column=4, value=person.get('call_time', '')).font = BODY_FONT
                ws.cell(row=row, column=5, value=person.get('phone', '')).font = BODY_FONT
                ws.cell(row=row, column=6, value=person.get('email', '')).font = BODY_FONT
                ws.cell(row=row, column=7, value=person.get('notes', '')).font = BODY_FONT

                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1


def _build_scenes_sheet(ws, scenes: List[Dict]):
    """Build the scenes sheet"""

    # Set column widths
    columns = [
        ('A', 'Scene #', 10),
        ('B', 'Int/Ext', 10),
        ('C', 'Day/Night', 10),
        ('D', 'Location', 25),
        ('E', 'Description', 40),
        ('F', 'Pages', 10),
        ('G', 'Notes', 30),
    ]

    for col_letter, _, width in columns:
        ws.column_dimensions[col_letter].width = width

    # Header row
    row = 1
    for idx, (_, header, _) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_BG
        cell.border = THIN_BORDER
    row += 1

    # Scene rows
    for idx, scene in enumerate(scenes):
        ws.cell(row=row, column=1, value=scene.get('scene_number', '')).font = BODY_FONT
        ws.cell(row=row, column=2, value=scene.get('int_ext', '')).font = BODY_FONT
        ws.cell(row=row, column=3, value=scene.get('day_night', '')).font = BODY_FONT
        ws.cell(row=row, column=4, value=scene.get('location', '')).font = BODY_FONT

        desc_cell = ws.cell(row=row, column=5, value=scene.get('description', ''))
        desc_cell.font = BODY_FONT
        desc_cell.alignment = Alignment(wrap_text=True)

        ws.cell(row=row, column=6, value=scene.get('page_count', '')).font = BODY_FONT

        notes_cell = ws.cell(row=row, column=7, value=scene.get('notes', ''))
        notes_cell.font = BODY_FONT
        notes_cell.alignment = Alignment(wrap_text=True)

        # Alternating row colors
        if idx % 2 == 1:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = ALT_ROW_BG

        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1


def _build_locations_sheet(ws, locations: List[Dict]):
    """Build the locations sheet"""

    # Set column widths
    columns = [
        ('A', 'Location Name', 25),
        ('B', 'Address', 35),
        ('C', 'Contact', 20),
        ('D', 'Phone', 15),
        ('E', 'Parking', 20),
        ('F', 'Notes', 30),
    ]

    for col_letter, _, width in columns:
        ws.column_dimensions[col_letter].width = width

    # Header row
    row = 1
    for idx, (_, header, _) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_BG
        cell.border = THIN_BORDER
    row += 1

    # Location rows
    for idx, location in enumerate(locations):
        ws.cell(row=row, column=1, value=location.get('name', '')).font = BODY_FONT

        addr_cell = ws.cell(row=row, column=2, value=location.get('address', ''))
        addr_cell.font = BODY_FONT
        addr_cell.alignment = Alignment(wrap_text=True)

        ws.cell(row=row, column=3, value=location.get('contact_name', '')).font = BODY_FONT
        ws.cell(row=row, column=4, value=location.get('contact_phone', '')).font = BODY_FONT
        ws.cell(row=row, column=5, value=location.get('parking_info', '')).font = BODY_FONT

        notes_cell = ws.cell(row=row, column=6, value=location.get('notes', ''))
        notes_cell.font = BODY_FONT
        notes_cell.alignment = Alignment(wrap_text=True)

        # Alternating row colors
        if idx % 2 == 1:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = ALT_ROW_BG

        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1
