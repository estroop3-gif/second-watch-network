"""
Episode Management API Endpoints
Handles episodes, seasons, approvals, and all episode-related CRUD operations
"""
from fastapi import APIRouter, HTTPException, Header, Query, UploadFile, File
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, date
import csv
import io
import json

from app.core.database import get_client, execute_query, execute_single

router = APIRouter()

# Valid enum values
PIPELINE_STAGES = ['DEVELOPMENT', 'PRE_PRO', 'PRODUCTION', 'POST', 'DELIVERED', 'RELEASED']
EDIT_STATUSES = ['NOT_STARTED', 'INGEST', 'ASSEMBLY', 'ROUGH_CUT', 'FINE_CUT', 'PICTURE_LOCK', 'SOUND', 'COLOR', 'MASTERING']
DELIVERY_STATUSES = ['NOT_STARTED', 'QC', 'CAPTIONS', 'ARTWORK', 'EXPORTS', 'DELIVERED', 'RELEASED']
SUBJECT_TYPES = ['CAST', 'CREW', 'CONTRIBUTOR', 'OTHER']
LIST_ITEM_KINDS = ['INTERVIEW', 'SCENE', 'SEGMENT']
DELIVERABLE_STATUSES = ['NOT_STARTED', 'IN_PROGRESS', 'READY_FOR_REVIEW', 'APPROVED', 'DELIVERED']
APPROVAL_TYPES = ['EDIT_LOCK', 'DELIVERY_APPROVAL']
APPROVAL_STATUSES = ['PENDING', 'APPROVED', 'REJECTED']


# =====================================================
# Pydantic Models
# =====================================================

class SeasonCreate(BaseModel):
    season_number: int = Field(..., ge=1)
    title: Optional[str] = None


class SeasonUpdate(BaseModel):
    season_number: Optional[int] = None
    title: Optional[str] = None


class EpisodeCreate(BaseModel):
    season_id: Optional[str] = None
    episode_number: int = Field(..., ge=1)
    episode_code: Optional[str] = None
    title: str = Field(..., min_length=1)
    logline: Optional[str] = None
    synopsis: Optional[str] = None
    planned_runtime_minutes: Optional[int] = None


class EpisodeUpdate(BaseModel):
    season_id: Optional[str] = None
    episode_number: Optional[int] = None
    episode_code: Optional[str] = None
    title: Optional[str] = None
    logline: Optional[str] = None
    synopsis: Optional[str] = None
    outline: Optional[str] = None
    beat_sheet: Optional[str] = None
    notes: Optional[str] = None
    pipeline_stage: Optional[str] = None
    edit_status: Optional[str] = None
    delivery_status: Optional[str] = None
    editor_user_id: Optional[str] = None
    ae_user_id: Optional[str] = None
    post_supervisor_user_id: Optional[str] = None
    planned_runtime_minutes: Optional[int] = None
    actual_runtime_minutes: Optional[int] = None


class SubjectCreate(BaseModel):
    subject_type: str
    name: str = Field(..., min_length=1)
    role: Optional[str] = None
    contact_info: Optional[str] = None
    notes: Optional[str] = None


class SubjectUpdate(BaseModel):
    subject_type: Optional[str] = None
    name: Optional[str] = None
    role: Optional[str] = None
    contact_info: Optional[str] = None
    notes: Optional[str] = None


class LocationCreate(BaseModel):
    name: str = Field(..., min_length=1)
    address: Optional[str] = None
    notes: Optional[str] = None


class LocationUpdate(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None


class ListItemCreate(BaseModel):
    kind: str
    title: str = Field(..., min_length=1)
    description: Optional[str] = None
    status: Optional[str] = None


class ListItemUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None


class ListItemReorder(BaseModel):
    item_id: str
    direction: str = Field(..., pattern='^(up|down)$')


class MilestoneCreate(BaseModel):
    milestone_type: str = Field(..., min_length=1)
    date: date
    notes: Optional[str] = None


class MilestoneUpdate(BaseModel):
    milestone_type: Optional[str] = None
    date: Optional[date] = None
    notes: Optional[str] = None


class DeliverableCreate(BaseModel):
    deliverable_type: str = Field(..., min_length=1)
    status: Optional[str] = 'NOT_STARTED'
    due_date: Optional[date] = None
    owner_user_id: Optional[str] = None
    notes: Optional[str] = None


class DeliverableUpdate(BaseModel):
    deliverable_type: Optional[str] = None
    status: Optional[str] = None
    due_date: Optional[date] = None
    owner_user_id: Optional[str] = None
    notes: Optional[str] = None


class DeliverableTemplateCreate(BaseModel):
    name: str = Field(..., min_length=1)
    items: List[Dict[str, Any]] = []


class ApplyTemplateRequest(BaseModel):
    template_id: str


class AssetLinkCreate(BaseModel):
    label: str = Field(..., min_length=1)
    url: str = Field(..., min_length=1)


class ShootDayCreate(BaseModel):
    production_day_id: str


class ApprovalRequest(BaseModel):
    approval_type: str
    notes: Optional[str] = None


class ApprovalDecision(BaseModel):
    decision: str = Field(..., pattern='^(APPROVE|REJECT)$')
    notes: Optional[str] = None


class StoryboardLink(BaseModel):
    storyboard_id: str


class EpisodeSettingsUpdate(BaseModel):
    settings_json: Dict[str, Any]


# =====================================================
# Helper Functions
# =====================================================

async def get_current_user_from_token(authorization: str = Header(None)) -> Dict[str, Any]:
    """Extract and validate user from Bearer token."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing or invalid authorization header")

    token = authorization.replace("Bearer ", "")

    # Decode JWT to get user info
    import jwt
    try:
        # Decode without verification for user info (Cognito already validated)
        decoded = jwt.decode(token, options={"verify_signature": False})
        cognito_id = decoded.get("sub")
        if not cognito_id:
            raise HTTPException(status_code=401, detail="Invalid token")

        # Get profile from cognito_user_id
        profile = execute_single(
            "SELECT id, email, display_name FROM profiles WHERE cognito_user_id = :cognito_id",
            {"cognito_id": cognito_id}
        )
        if not profile:
            raise HTTPException(status_code=401, detail="User profile not found")

        return {
            "user_id": str(profile["id"]),
            "email": profile["email"],
            "display_name": profile.get("display_name"),
            "cognito_id": cognito_id
        }
    except jwt.DecodeError:
        raise HTTPException(status_code=401, detail="Invalid token format")


async def verify_project_access(project_id: str, user_id: str) -> Dict[str, Any]:
    """Verify user has access to project and return project data."""
    project = execute_single(
        """SELECT p.*, pm.role as member_role
           FROM backlot_projects p
           LEFT JOIN backlot_project_members pm ON pm.project_id = p.id AND pm.user_id = :user_id
           WHERE p.id = :project_id""",
        {"project_id": project_id, "user_id": user_id}
    )

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Check if user is owner or member
    is_owner = str(project.get("owner_id")) == user_id
    is_member = project.get("member_role") is not None

    if not is_owner and not is_member:
        raise HTTPException(status_code=403, detail="Access denied to this project")

    return {
        **dict(project),
        "is_owner": is_owner,
        "user_role": "OWNER" if is_owner else project.get("member_role", "MEMBER")
    }


async def get_episode_settings(project_id: str) -> Dict[str, Any]:
    """Get episode settings for project, creating defaults if needed."""
    settings = execute_single(
        "SELECT settings_json FROM project_episode_settings WHERE project_id = :project_id",
        {"project_id": project_id}
    )

    if settings:
        return settings["settings_json"]

    # Return defaults
    return {
        "canCreate": ["OWNER", "ADMIN", "PRODUCER"],
        "canEdit": ["OWNER", "ADMIN", "PRODUCER", "COORDINATOR"],
        "canDelete": ["OWNER", "ADMIN"],
        "canApproveEditLock": ["OWNER", "ADMIN", "SHOWRUNNER"],
        "canApproveDelivery": ["OWNER", "ADMIN", "DISTRIBUTION"],
        "defaultAssignees": {"editorUserId": None, "postSupervisorUserId": None}
    }


def check_permission(user_role: str, allowed_roles: List[str]) -> bool:
    """Check if user role is in allowed roles."""
    return user_role in allowed_roles


async def verify_episode_access(project_id: str, episode_id: str, user_id: str, require_edit: bool = False) -> Dict[str, Any]:
    """Verify episode exists and user has access."""
    project = await verify_project_access(project_id, user_id)
    settings = await get_episode_settings(project_id)

    episode = execute_single(
        "SELECT * FROM episodes WHERE id = :episode_id AND project_id = :project_id",
        {"episode_id": episode_id, "project_id": project_id}
    )

    if not episode:
        raise HTTPException(status_code=404, detail="Episode not found")

    if require_edit:
        # Check if episode is locked
        if episode.get("is_edit_locked"):
            # Only users with approve permission can edit locked episodes
            can_approve = check_permission(project["user_role"], settings.get("canApproveEditLock", ["OWNER"]))
            if not can_approve:
                raise HTTPException(status_code=403, detail="Episode is locked for editing")

        # Check edit permission
        if not check_permission(project["user_role"], settings.get("canEdit", ["OWNER"])):
            raise HTTPException(status_code=403, detail="No permission to edit episodes")

    return {"episode": dict(episode), "project": project, "settings": settings}


# =====================================================
# Season Endpoints
# =====================================================

@router.get("/projects/{project_id}/seasons")
async def list_seasons(
    project_id: str,
    authorization: str = Header(None)
):
    """List all seasons for a project."""
    user = await get_current_user_from_token(authorization)
    await verify_project_access(project_id, user["user_id"])

    seasons = execute_query(
        """SELECT s.*,
              (SELECT COUNT(*) FROM episodes e WHERE e.season_id = s.id) as episode_count
           FROM seasons s
           WHERE s.project_id = :project_id
           ORDER BY s.season_number""",
        {"project_id": project_id}
    )

    return {"seasons": [dict(s) for s in seasons]}


@router.post("/projects/{project_id}/seasons")
async def create_season(
    project_id: str,
    data: SeasonCreate,
    authorization: str = Header(None)
):
    """Create a new season."""
    user = await get_current_user_from_token(authorization)
    project = await verify_project_access(project_id, user["user_id"])
    settings = await get_episode_settings(project_id)

    if not check_permission(project["user_role"], settings.get("canCreate", ["OWNER"])):
        raise HTTPException(status_code=403, detail="No permission to create seasons")

    client = get_client()
    result = client.table("seasons").insert({
        "project_id": project_id,
        "season_number": data.season_number,
        "title": data.title
    }).execute()

    return result.data[0] if result.data else {"success": True}


@router.put("/projects/{project_id}/seasons/{season_id}")
async def update_season(
    project_id: str,
    season_id: str,
    data: SeasonUpdate,
    authorization: str = Header(None)
):
    """Update a season."""
    user = await get_current_user_from_token(authorization)
    project = await verify_project_access(project_id, user["user_id"])
    settings = await get_episode_settings(project_id)

    if not check_permission(project["user_role"], settings.get("canEdit", ["OWNER"])):
        raise HTTPException(status_code=403, detail="No permission to edit seasons")

    update_data = {k: v for k, v in data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow().isoformat()

    client = get_client()
    result = client.table("seasons").update(update_data).eq("id", season_id).eq("project_id", project_id).execute()

    if not result.data:
        raise HTTPException(status_code=404, detail="Season not found")

    return result.data[0]


@router.delete("/projects/{project_id}/seasons/{season_id}")
async def delete_season(
    project_id: str,
    season_id: str,
    authorization: str = Header(None)
):
    """Delete a season."""
    user = await get_current_user_from_token(authorization)
    project = await verify_project_access(project_id, user["user_id"])
    settings = await get_episode_settings(project_id)

    if not check_permission(project["user_role"], settings.get("canDelete", ["OWNER"])):
        raise HTTPException(status_code=403, detail="No permission to delete seasons")

    client = get_client()
    client.table("seasons").delete().eq("id", season_id).eq("project_id", project_id).execute()

    return {"success": True}


# =====================================================
# Episode List and CRUD
# =====================================================

@router.get("/projects/{project_id}/episodes")
async def list_episodes(
    project_id: str,
    season_id: Optional[str] = None,
    search: Optional[str] = None,
    pipeline_stage: Optional[str] = None,
    authorization: str = Header(None)
):
    """List episodes for a project with optional filters."""
    user = await get_current_user_from_token(authorization)
    await verify_project_access(project_id, user["user_id"])

    query = """
        SELECT e.*,
               s.season_number, s.title as season_title,
               (SELECT COUNT(*) FROM episode_approvals ea
                WHERE ea.episode_id = e.id AND ea.status = 'PENDING') as pending_approvals
        FROM episodes e
        LEFT JOIN seasons s ON s.id = e.season_id
        WHERE e.project_id = :project_id
    """
    params = {"project_id": project_id}

    if season_id:
        query += " AND e.season_id = :season_id"
        params["season_id"] = season_id

    if pipeline_stage:
        query += " AND e.pipeline_stage = :pipeline_stage"
        params["pipeline_stage"] = pipeline_stage

    if search:
        query += " AND (e.title ILIKE :search OR e.episode_code ILIKE :search)"
        params["search"] = f"%{search}%"

    query += " ORDER BY s.season_number NULLS LAST, e.episode_number"

    episodes = execute_query(query, params)

    return {"episodes": [dict(e) for e in episodes]}


@router.post("/projects/{project_id}/episodes")
async def create_episode(
    project_id: str,
    data: EpisodeCreate,
    authorization: str = Header(None)
):
    """Create a new episode."""
    user = await get_current_user_from_token(authorization)
    project = await verify_project_access(project_id, user["user_id"])
    settings = await get_episode_settings(project_id)

    if not check_permission(project["user_role"], settings.get("canCreate", ["OWNER"])):
        raise HTTPException(status_code=403, detail="No permission to create episodes")

    # Generate episode code if not provided
    episode_code = data.episode_code
    if not episode_code:
        if data.season_id:
            season = execute_single(
                "SELECT season_number FROM seasons WHERE id = :season_id",
                {"season_id": data.season_id}
            )
            if season:
                episode_code = f"S{season['season_number']:02d}E{data.episode_number:02d}"
            else:
                episode_code = f"E{data.episode_number:02d}"
        else:
            episode_code = f"E{data.episode_number:02d}"

    client = get_client()
    result = client.table("episodes").insert({
        "project_id": project_id,
        "season_id": data.season_id,
        "episode_number": data.episode_number,
        "episode_code": episode_code,
        "title": data.title,
        "logline": data.logline,
        "synopsis": data.synopsis,
        "planned_runtime_minutes": data.planned_runtime_minutes,
        "created_by_user_id": user["user_id"]
    }).execute()

    return result.data[0] if result.data else {"success": True}


@router.get("/projects/{project_id}/episodes/{episode_id}")
async def get_episode(
    project_id: str,
    episode_id: str,
    authorization: str = Header(None)
):
    """Get episode detail with all related data."""
    user = await get_current_user_from_token(authorization)
    await verify_project_access(project_id, user["user_id"])

    episode = execute_single(
        """SELECT e.*, s.season_number, s.title as season_title
           FROM episodes e
           LEFT JOIN seasons s ON s.id = e.season_id
           WHERE e.id = :episode_id AND e.project_id = :project_id""",
        {"episode_id": episode_id, "project_id": project_id}
    )

    if not episode:
        raise HTTPException(status_code=404, detail="Episode not found")

    # Get related data
    subjects = execute_query(
        "SELECT * FROM episode_subjects WHERE episode_id = :episode_id ORDER BY name",
        {"episode_id": episode_id}
    )

    locations = execute_query(
        "SELECT * FROM episode_locations WHERE episode_id = :episode_id ORDER BY name",
        {"episode_id": episode_id}
    )

    list_items = execute_query(
        "SELECT * FROM episode_list_items WHERE episode_id = :episode_id ORDER BY kind, sort_order",
        {"episode_id": episode_id}
    )

    milestones = execute_query(
        "SELECT * FROM episode_milestones WHERE episode_id = :episode_id ORDER BY date",
        {"episode_id": episode_id}
    )

    deliverables = execute_query(
        "SELECT * FROM episode_deliverables WHERE episode_id = :episode_id ORDER BY deliverable_type",
        {"episode_id": episode_id}
    )

    asset_links = execute_query(
        "SELECT * FROM episode_asset_links WHERE episode_id = :episode_id ORDER BY label",
        {"episode_id": episode_id}
    )

    shoot_days = execute_query(
        """SELECT esd.*, pd.date, pd.day_type, pd.title as day_title
           FROM episode_shoot_days esd
           JOIN backlot_production_days pd ON pd.id = esd.production_day_id
           WHERE esd.episode_id = :episode_id
           ORDER BY pd.date""",
        {"episode_id": episode_id}
    )

    approvals = execute_query(
        """SELECT ea.*,
                  req.display_name as requested_by_name,
                  dec.display_name as decided_by_name
           FROM episode_approvals ea
           LEFT JOIN profiles req ON req.id = ea.requested_by_user_id
           LEFT JOIN profiles dec ON dec.id = ea.decided_by_user_id
           WHERE ea.episode_id = :episode_id
           ORDER BY ea.requested_at DESC""",
        {"episode_id": episode_id}
    )

    storyboards = execute_query(
        "SELECT id, title, status, aspect_ratio FROM storyboards WHERE episode_id = :episode_id ORDER BY title",
        {"episode_id": episode_id}
    )

    return {
        **dict(episode),
        "subjects": [dict(s) for s in subjects],
        "locations": [dict(l) for l in locations],
        "list_items": [dict(i) for i in list_items],
        "milestones": [dict(m) for m in milestones],
        "deliverables": [dict(d) for d in deliverables],
        "asset_links": [dict(a) for a in asset_links],
        "shoot_days": [dict(sd) for sd in shoot_days],
        "approvals": [dict(ap) for ap in approvals],
        "storyboards": [dict(sb) for sb in storyboards]
    }


@router.put("/projects/{project_id}/episodes/{episode_id}")
async def update_episode(
    project_id: str,
    episode_id: str,
    data: EpisodeUpdate,
    authorization: str = Header(None)
):
    """Update an episode."""
    user = await get_current_user_from_token(authorization)
    access = await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    # Validate enum values
    if data.pipeline_stage and data.pipeline_stage not in PIPELINE_STAGES:
        raise HTTPException(status_code=400, detail=f"Invalid pipeline_stage. Must be one of: {PIPELINE_STAGES}")
    if data.edit_status and data.edit_status not in EDIT_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid edit_status. Must be one of: {EDIT_STATUSES}")
    if data.delivery_status and data.delivery_status not in DELIVERY_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid delivery_status. Must be one of: {DELIVERY_STATUSES}")

    update_data = {k: v for k, v in data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow().isoformat()

    client = get_client()
    result = client.table("episodes").update(update_data).eq("id", episode_id).eq("project_id", project_id).execute()

    if not result.data:
        raise HTTPException(status_code=404, detail="Episode not found")

    return result.data[0]


@router.delete("/projects/{project_id}/episodes/{episode_id}")
async def delete_episode(
    project_id: str,
    episode_id: str,
    authorization: str = Header(None)
):
    """Delete an episode."""
    user = await get_current_user_from_token(authorization)
    project = await verify_project_access(project_id, user["user_id"])
    settings = await get_episode_settings(project_id)

    if not check_permission(project["user_role"], settings.get("canDelete", ["OWNER"])):
        raise HTTPException(status_code=403, detail="No permission to delete episodes")

    client = get_client()
    client.table("episodes").delete().eq("id", episode_id).eq("project_id", project_id).execute()

    return {"success": True}


# =====================================================
# Subjects CRUD
# =====================================================

@router.post("/projects/{project_id}/episodes/{episode_id}/subjects")
async def create_subject(
    project_id: str,
    episode_id: str,
    data: SubjectCreate,
    authorization: str = Header(None)
):
    """Add a subject to an episode."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    if data.subject_type not in SUBJECT_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid subject_type. Must be one of: {SUBJECT_TYPES}")

    client = get_client()
    result = client.table("episode_subjects").insert({
        "episode_id": episode_id,
        "subject_type": data.subject_type,
        "name": data.name,
        "role": data.role,
        "contact_info": data.contact_info,
        "notes": data.notes
    }).execute()

    return result.data[0] if result.data else {"success": True}


@router.put("/projects/{project_id}/episodes/{episode_id}/subjects/{subject_id}")
async def update_subject(
    project_id: str,
    episode_id: str,
    subject_id: str,
    data: SubjectUpdate,
    authorization: str = Header(None)
):
    """Update a subject."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    if data.subject_type and data.subject_type not in SUBJECT_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid subject_type. Must be one of: {SUBJECT_TYPES}")

    update_data = {k: v for k, v in data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow().isoformat()

    client = get_client()
    result = client.table("episode_subjects").update(update_data).eq("id", subject_id).eq("episode_id", episode_id).execute()

    return result.data[0] if result.data else {"success": True}


@router.delete("/projects/{project_id}/episodes/{episode_id}/subjects/{subject_id}")
async def delete_subject(
    project_id: str,
    episode_id: str,
    subject_id: str,
    authorization: str = Header(None)
):
    """Delete a subject."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    client = get_client()
    client.table("episode_subjects").delete().eq("id", subject_id).eq("episode_id", episode_id).execute()

    return {"success": True}


# =====================================================
# Locations CRUD
# =====================================================

@router.post("/projects/{project_id}/episodes/{episode_id}/locations")
async def create_location(
    project_id: str,
    episode_id: str,
    data: LocationCreate,
    authorization: str = Header(None)
):
    """Add a location to an episode."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    client = get_client()
    result = client.table("episode_locations").insert({
        "episode_id": episode_id,
        "name": data.name,
        "address": data.address,
        "notes": data.notes
    }).execute()

    return result.data[0] if result.data else {"success": True}


@router.put("/projects/{project_id}/episodes/{episode_id}/locations/{location_id}")
async def update_location(
    project_id: str,
    episode_id: str,
    location_id: str,
    data: LocationUpdate,
    authorization: str = Header(None)
):
    """Update a location."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    update_data = {k: v for k, v in data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow().isoformat()

    client = get_client()
    result = client.table("episode_locations").update(update_data).eq("id", location_id).eq("episode_id", episode_id).execute()

    return result.data[0] if result.data else {"success": True}


@router.delete("/projects/{project_id}/episodes/{episode_id}/locations/{location_id}")
async def delete_location(
    project_id: str,
    episode_id: str,
    location_id: str,
    authorization: str = Header(None)
):
    """Delete a location."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    client = get_client()
    client.table("episode_locations").delete().eq("id", location_id).eq("episode_id", episode_id).execute()

    return {"success": True}


# =====================================================
# List Items CRUD (Interviews, Scenes, Segments)
# =====================================================

@router.post("/projects/{project_id}/episodes/{episode_id}/list-items")
async def create_list_item(
    project_id: str,
    episode_id: str,
    data: ListItemCreate,
    authorization: str = Header(None)
):
    """Add a list item to an episode."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    if data.kind not in LIST_ITEM_KINDS:
        raise HTTPException(status_code=400, detail=f"Invalid kind. Must be one of: {LIST_ITEM_KINDS}")

    # Get max sort_order for this kind
    max_order = execute_single(
        "SELECT COALESCE(MAX(sort_order), -1) as max_order FROM episode_list_items WHERE episode_id = :episode_id AND kind = :kind",
        {"episode_id": episode_id, "kind": data.kind}
    )
    next_order = (max_order["max_order"] or -1) + 1

    client = get_client()
    result = client.table("episode_list_items").insert({
        "episode_id": episode_id,
        "kind": data.kind,
        "sort_order": next_order,
        "title": data.title,
        "description": data.description,
        "status": data.status
    }).execute()

    return result.data[0] if result.data else {"success": True}


@router.put("/projects/{project_id}/episodes/{episode_id}/list-items/{item_id}")
async def update_list_item(
    project_id: str,
    episode_id: str,
    item_id: str,
    data: ListItemUpdate,
    authorization: str = Header(None)
):
    """Update a list item."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    update_data = {k: v for k, v in data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow().isoformat()

    client = get_client()
    result = client.table("episode_list_items").update(update_data).eq("id", item_id).eq("episode_id", episode_id).execute()

    return result.data[0] if result.data else {"success": True}


@router.delete("/projects/{project_id}/episodes/{episode_id}/list-items/{item_id}")
async def delete_list_item(
    project_id: str,
    episode_id: str,
    item_id: str,
    authorization: str = Header(None)
):
    """Delete a list item."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    client = get_client()
    client.table("episode_list_items").delete().eq("id", item_id).eq("episode_id", episode_id).execute()

    return {"success": True}


@router.post("/projects/{project_id}/episodes/{episode_id}/list-items/reorder")
async def reorder_list_item(
    project_id: str,
    episode_id: str,
    data: ListItemReorder,
    authorization: str = Header(None)
):
    """Reorder a list item up or down within its kind."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    # Get current item
    item = execute_single(
        "SELECT * FROM episode_list_items WHERE id = :item_id AND episode_id = :episode_id",
        {"item_id": data.item_id, "episode_id": episode_id}
    )
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    # Get adjacent item
    if data.direction == "up":
        adjacent = execute_single(
            """SELECT * FROM episode_list_items
               WHERE episode_id = :episode_id AND kind = :kind AND sort_order < :sort_order
               ORDER BY sort_order DESC LIMIT 1""",
            {"episode_id": episode_id, "kind": item["kind"], "sort_order": item["sort_order"]}
        )
    else:
        adjacent = execute_single(
            """SELECT * FROM episode_list_items
               WHERE episode_id = :episode_id AND kind = :kind AND sort_order > :sort_order
               ORDER BY sort_order ASC LIMIT 1""",
            {"episode_id": episode_id, "kind": item["kind"], "sort_order": item["sort_order"]}
        )

    if not adjacent:
        return {"success": True, "message": "Already at boundary"}

    # Swap sort_orders
    client = get_client()
    # Use temporary value to avoid unique constraint
    client.table("episode_list_items").update({"sort_order": -999}).eq("id", data.item_id).execute()
    client.table("episode_list_items").update({"sort_order": item["sort_order"]}).eq("id", adjacent["id"]).execute()
    client.table("episode_list_items").update({"sort_order": adjacent["sort_order"]}).eq("id", data.item_id).execute()

    return {"success": True}


# =====================================================
# Milestones CRUD
# =====================================================

@router.post("/projects/{project_id}/episodes/{episode_id}/milestones")
async def create_milestone(
    project_id: str,
    episode_id: str,
    data: MilestoneCreate,
    authorization: str = Header(None)
):
    """Add a milestone to an episode."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    client = get_client()
    result = client.table("episode_milestones").insert({
        "episode_id": episode_id,
        "milestone_type": data.milestone_type,
        "date": data.date.isoformat(),
        "notes": data.notes
    }).execute()

    return result.data[0] if result.data else {"success": True}


@router.put("/projects/{project_id}/episodes/{episode_id}/milestones/{milestone_id}")
async def update_milestone(
    project_id: str,
    episode_id: str,
    milestone_id: str,
    data: MilestoneUpdate,
    authorization: str = Header(None)
):
    """Update a milestone."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    update_data = {}
    if data.milestone_type is not None:
        update_data["milestone_type"] = data.milestone_type
    if data.date is not None:
        update_data["date"] = data.date.isoformat()
    if data.notes is not None:
        update_data["notes"] = data.notes
    update_data["updated_at"] = datetime.utcnow().isoformat()

    client = get_client()
    result = client.table("episode_milestones").update(update_data).eq("id", milestone_id).eq("episode_id", episode_id).execute()

    return result.data[0] if result.data else {"success": True}


@router.delete("/projects/{project_id}/episodes/{episode_id}/milestones/{milestone_id}")
async def delete_milestone(
    project_id: str,
    episode_id: str,
    milestone_id: str,
    authorization: str = Header(None)
):
    """Delete a milestone."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    client = get_client()
    client.table("episode_milestones").delete().eq("id", milestone_id).eq("episode_id", episode_id).execute()

    return {"success": True}


# =====================================================
# Deliverables CRUD
# =====================================================

@router.get("/projects/{project_id}/deliverable-templates")
async def list_deliverable_templates(
    project_id: str,
    authorization: str = Header(None)
):
    """List deliverable templates for a project."""
    user = await get_current_user_from_token(authorization)
    await verify_project_access(project_id, user["user_id"])

    templates = execute_query(
        "SELECT * FROM project_deliverable_templates WHERE project_id = :project_id ORDER BY name",
        {"project_id": project_id}
    )

    return {"templates": [dict(t) for t in templates]}


@router.post("/projects/{project_id}/deliverable-templates")
async def create_deliverable_template(
    project_id: str,
    data: DeliverableTemplateCreate,
    authorization: str = Header(None)
):
    """Create a deliverable template."""
    user = await get_current_user_from_token(authorization)
    project = await verify_project_access(project_id, user["user_id"])
    settings = await get_episode_settings(project_id)

    if not check_permission(project["user_role"], settings.get("canCreate", ["OWNER"])):
        raise HTTPException(status_code=403, detail="No permission to create templates")

    client = get_client()
    result = client.table("project_deliverable_templates").insert({
        "project_id": project_id,
        "name": data.name,
        "items": json.dumps(data.items)
    }).execute()

    return result.data[0] if result.data else {"success": True}


@router.post("/projects/{project_id}/episodes/{episode_id}/deliverables")
async def create_deliverable(
    project_id: str,
    episode_id: str,
    data: DeliverableCreate,
    authorization: str = Header(None)
):
    """Add a deliverable to an episode."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    if data.status and data.status not in DELIVERABLE_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {DELIVERABLE_STATUSES}")

    insert_data = {
        "episode_id": episode_id,
        "deliverable_type": data.deliverable_type,
        "status": data.status or "NOT_STARTED",
        "notes": data.notes
    }
    if data.due_date:
        insert_data["due_date"] = data.due_date.isoformat()
    if data.owner_user_id:
        insert_data["owner_user_id"] = data.owner_user_id

    client = get_client()
    result = client.table("episode_deliverables").insert(insert_data).execute()

    return result.data[0] if result.data else {"success": True}


@router.put("/projects/{project_id}/episodes/{episode_id}/deliverables/{deliverable_id}")
async def update_deliverable(
    project_id: str,
    episode_id: str,
    deliverable_id: str,
    data: DeliverableUpdate,
    authorization: str = Header(None)
):
    """Update a deliverable."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    if data.status and data.status not in DELIVERABLE_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {DELIVERABLE_STATUSES}")

    update_data = {}
    if data.deliverable_type is not None:
        update_data["deliverable_type"] = data.deliverable_type
    if data.status is not None:
        update_data["status"] = data.status
    if data.due_date is not None:
        update_data["due_date"] = data.due_date.isoformat()
    if data.owner_user_id is not None:
        update_data["owner_user_id"] = data.owner_user_id
    if data.notes is not None:
        update_data["notes"] = data.notes
    update_data["updated_at"] = datetime.utcnow().isoformat()

    client = get_client()
    result = client.table("episode_deliverables").update(update_data).eq("id", deliverable_id).eq("episode_id", episode_id).execute()

    return result.data[0] if result.data else {"success": True}


@router.delete("/projects/{project_id}/episodes/{episode_id}/deliverables/{deliverable_id}")
async def delete_deliverable(
    project_id: str,
    episode_id: str,
    deliverable_id: str,
    authorization: str = Header(None)
):
    """Delete a deliverable."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    client = get_client()
    client.table("episode_deliverables").delete().eq("id", deliverable_id).eq("episode_id", episode_id).execute()

    return {"success": True}


@router.post("/projects/{project_id}/episodes/{episode_id}/deliverables/apply-template")
async def apply_deliverable_template(
    project_id: str,
    episode_id: str,
    data: ApplyTemplateRequest,
    authorization: str = Header(None)
):
    """Apply a deliverable template to an episode."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    template = execute_single(
        "SELECT * FROM project_deliverable_templates WHERE id = :template_id AND project_id = :project_id",
        {"template_id": data.template_id, "project_id": project_id}
    )

    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    items = template.get("items", [])
    if isinstance(items, str):
        items = json.loads(items)

    client = get_client()
    created = []
    for item in items:
        result = client.table("episode_deliverables").insert({
            "episode_id": episode_id,
            "deliverable_type": item.get("deliverable_type", item.get("type", "Unknown")),
            "status": "NOT_STARTED",
            "notes": item.get("notes")
        }).execute()
        if result.data:
            created.append(result.data[0])

    return {"success": True, "created": len(created)}


# =====================================================
# Asset Links CRUD
# =====================================================

@router.post("/projects/{project_id}/episodes/{episode_id}/asset-links")
async def create_asset_link(
    project_id: str,
    episode_id: str,
    data: AssetLinkCreate,
    authorization: str = Header(None)
):
    """Add an asset link to an episode."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    client = get_client()
    result = client.table("episode_asset_links").insert({
        "episode_id": episode_id,
        "label": data.label,
        "url": data.url
    }).execute()

    return result.data[0] if result.data else {"success": True}


@router.delete("/projects/{project_id}/episodes/{episode_id}/asset-links/{link_id}")
async def delete_asset_link(
    project_id: str,
    episode_id: str,
    link_id: str,
    authorization: str = Header(None)
):
    """Delete an asset link."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    client = get_client()
    client.table("episode_asset_links").delete().eq("id", link_id).eq("episode_id", episode_id).execute()

    return {"success": True}


# =====================================================
# Shoot Days (DOOD Integration)
# =====================================================

@router.get("/projects/{project_id}/project-days")
async def list_project_days(
    project_id: str,
    start: Optional[str] = None,
    end: Optional[str] = None,
    authorization: str = Header(None)
):
    """List production days for a project (for shoot day tagging)."""
    user = await get_current_user_from_token(authorization)
    await verify_project_access(project_id, user["user_id"])

    query = "SELECT * FROM backlot_production_days WHERE project_id = :project_id"
    params = {"project_id": project_id}

    if start:
        query += " AND date >= :start"
        params["start"] = start
    if end:
        query += " AND date <= :end"
        params["end"] = end

    query += " ORDER BY date"

    days = execute_query(query, params)

    return {"days": [dict(d) for d in days]}


@router.post("/projects/{project_id}/episodes/{episode_id}/shoot-days")
async def tag_shoot_day(
    project_id: str,
    episode_id: str,
    data: ShootDayCreate,
    authorization: str = Header(None)
):
    """Tag a production day as a shoot day for this episode."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    # Verify production day belongs to project
    day = execute_single(
        "SELECT id FROM backlot_production_days WHERE id = :day_id AND project_id = :project_id",
        {"day_id": data.production_day_id, "project_id": project_id}
    )
    if not day:
        raise HTTPException(status_code=404, detail="Production day not found")

    client = get_client()
    try:
        result = client.table("episode_shoot_days").insert({
            "episode_id": episode_id,
            "production_day_id": data.production_day_id
        }).execute()
        return result.data[0] if result.data else {"success": True}
    except Exception as e:
        if "unique" in str(e).lower():
            raise HTTPException(status_code=409, detail="Day already tagged")
        raise


@router.delete("/projects/{project_id}/episodes/{episode_id}/shoot-days/{shoot_day_id}")
async def untag_shoot_day(
    project_id: str,
    episode_id: str,
    shoot_day_id: str,
    authorization: str = Header(None)
):
    """Remove a shoot day tag from an episode."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    client = get_client()
    client.table("episode_shoot_days").delete().eq("id", shoot_day_id).eq("episode_id", episode_id).execute()

    return {"success": True}


# =====================================================
# Storyboard Linking
# =====================================================

@router.get("/projects/{project_id}/storyboards")
async def list_project_storyboards(
    project_id: str,
    unlinked_only: bool = False,
    authorization: str = Header(None)
):
    """List storyboards for a project (optionally only unlinked ones)."""
    user = await get_current_user_from_token(authorization)
    await verify_project_access(project_id, user["user_id"])

    query = "SELECT id, title, status, aspect_ratio, episode_id FROM storyboards WHERE project_id = :project_id"
    params = {"project_id": project_id}

    if unlinked_only:
        query += " AND episode_id IS NULL"

    query += " ORDER BY title"

    storyboards = execute_query(query, params)

    return {"storyboards": [dict(sb) for sb in storyboards]}


@router.post("/projects/{project_id}/episodes/{episode_id}/link-storyboard")
async def link_storyboard(
    project_id: str,
    episode_id: str,
    data: StoryboardLink,
    authorization: str = Header(None)
):
    """Link a storyboard to an episode."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    # Verify storyboard belongs to project
    sb = execute_single(
        "SELECT id FROM storyboards WHERE id = :sb_id AND project_id = :project_id",
        {"sb_id": data.storyboard_id, "project_id": project_id}
    )
    if not sb:
        raise HTTPException(status_code=404, detail="Storyboard not found")

    client = get_client()
    result = client.table("storyboards").update({
        "episode_id": episode_id,
        "updated_at": datetime.utcnow().isoformat()
    }).eq("id", data.storyboard_id).execute()

    return result.data[0] if result.data else {"success": True}


@router.post("/projects/{project_id}/episodes/{episode_id}/unlink-storyboard")
async def unlink_storyboard(
    project_id: str,
    episode_id: str,
    data: StoryboardLink,
    authorization: str = Header(None)
):
    """Unlink a storyboard from an episode."""
    user = await get_current_user_from_token(authorization)
    await verify_episode_access(project_id, episode_id, user["user_id"], require_edit=True)

    client = get_client()
    result = client.table("storyboards").update({
        "episode_id": None,
        "updated_at": datetime.utcnow().isoformat()
    }).eq("id", data.storyboard_id).eq("episode_id", episode_id).execute()

    return {"success": True}


# =====================================================
# Approvals
# =====================================================

@router.post("/projects/{project_id}/episodes/{episode_id}/approvals")
async def request_approval(
    project_id: str,
    episode_id: str,
    data: ApprovalRequest,
    authorization: str = Header(None)
):
    """Request an approval (edit lock or delivery approval)."""
    user = await get_current_user_from_token(authorization)
    project = await verify_project_access(project_id, user["user_id"])
    settings = await get_episode_settings(project_id)

    # Verify edit permission to request approval
    if not check_permission(project["user_role"], settings.get("canEdit", ["OWNER"])):
        raise HTTPException(status_code=403, detail="No permission to request approvals")

    if data.approval_type not in APPROVAL_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid approval_type. Must be one of: {APPROVAL_TYPES}")

    # Verify episode exists
    episode = execute_single(
        "SELECT id FROM episodes WHERE id = :episode_id AND project_id = :project_id",
        {"episode_id": episode_id, "project_id": project_id}
    )
    if not episode:
        raise HTTPException(status_code=404, detail="Episode not found")

    client = get_client()
    result = client.table("episode_approvals").insert({
        "episode_id": episode_id,
        "approval_type": data.approval_type,
        "status": "PENDING",
        "requested_by_user_id": user["user_id"],
        "notes": data.notes
    }).execute()

    return result.data[0] if result.data else {"success": True}


@router.post("/projects/{project_id}/episodes/{episode_id}/approvals/{approval_id}/decide")
async def decide_approval(
    project_id: str,
    episode_id: str,
    approval_id: str,
    data: ApprovalDecision,
    authorization: str = Header(None)
):
    """Approve or reject an approval request."""
    user = await get_current_user_from_token(authorization)
    project = await verify_project_access(project_id, user["user_id"])
    settings = await get_episode_settings(project_id)

    # Get approval
    approval = execute_single(
        "SELECT * FROM episode_approvals WHERE id = :approval_id AND episode_id = :episode_id",
        {"approval_id": approval_id, "episode_id": episode_id}
    )
    if not approval:
        raise HTTPException(status_code=404, detail="Approval not found")

    if approval["status"] != "PENDING":
        raise HTTPException(status_code=400, detail="Approval already decided")

    # Check permission based on approval type
    if approval["approval_type"] == "EDIT_LOCK":
        allowed = settings.get("canApproveEditLock", ["OWNER"])
    else:
        allowed = settings.get("canApproveDelivery", ["OWNER"])

    if not check_permission(project["user_role"], allowed):
        raise HTTPException(status_code=403, detail="No permission to decide this approval")

    new_status = "APPROVED" if data.decision == "APPROVE" else "REJECTED"

    client = get_client()

    # Update approval
    client.table("episode_approvals").update({
        "status": new_status,
        "decided_by_user_id": user["user_id"],
        "decided_at": datetime.utcnow().isoformat(),
        "notes": data.notes or approval.get("notes")
    }).eq("id", approval_id).execute()

    # If approving edit lock, set is_edit_locked on episode
    if approval["approval_type"] == "EDIT_LOCK" and new_status == "APPROVED":
        client.table("episodes").update({
            "is_edit_locked": True,
            "updated_at": datetime.utcnow().isoformat()
        }).eq("id", episode_id).execute()

    return {"success": True, "status": new_status}


@router.post("/projects/{project_id}/episodes/{episode_id}/unlock")
async def unlock_episode(
    project_id: str,
    episode_id: str,
    authorization: str = Header(None)
):
    """Unlock an episode (creates approval record and sets is_edit_locked false)."""
    user = await get_current_user_from_token(authorization)
    project = await verify_project_access(project_id, user["user_id"])
    settings = await get_episode_settings(project_id)

    # Check permission
    if not check_permission(project["user_role"], settings.get("canApproveEditLock", ["OWNER"])):
        raise HTTPException(status_code=403, detail="No permission to unlock episodes")

    client = get_client()

    # Create approval record for unlock
    client.table("episode_approvals").insert({
        "episode_id": episode_id,
        "approval_type": "EDIT_LOCK",
        "status": "APPROVED",
        "requested_by_user_id": user["user_id"],
        "decided_by_user_id": user["user_id"],
        "decided_at": datetime.utcnow().isoformat(),
        "notes": "Unlocked by authorized user"
    }).execute()

    # Set is_edit_locked false
    client.table("episodes").update({
        "is_edit_locked": False,
        "updated_at": datetime.utcnow().isoformat()
    }).eq("id", episode_id).eq("project_id", project_id).execute()

    return {"success": True}


# =====================================================
# Settings
# =====================================================

@router.get("/projects/{project_id}/episode-settings")
async def get_episode_settings_endpoint(
    project_id: str,
    authorization: str = Header(None)
):
    """Get episode settings for a project."""
    user = await get_current_user_from_token(authorization)
    await verify_project_access(project_id, user["user_id"])

    settings = await get_episode_settings(project_id)
    return {"settings": settings}


@router.put("/projects/{project_id}/episode-settings")
async def update_episode_settings(
    project_id: str,
    data: EpisodeSettingsUpdate,
    authorization: str = Header(None)
):
    """Update episode settings for a project."""
    user = await get_current_user_from_token(authorization)
    project = await verify_project_access(project_id, user["user_id"])

    # Only owner can update settings
    if not project["is_owner"]:
        raise HTTPException(status_code=403, detail="Only project owner can update episode settings")

    client = get_client()

    # Upsert settings
    existing = execute_single(
        "SELECT id FROM project_episode_settings WHERE project_id = :project_id",
        {"project_id": project_id}
    )

    if existing:
        result = client.table("project_episode_settings").update({
            "settings_json": json.dumps(data.settings_json),
            "updated_at": datetime.utcnow().isoformat()
        }).eq("project_id", project_id).execute()
    else:
        result = client.table("project_episode_settings").insert({
            "project_id": project_id,
            "settings_json": json.dumps(data.settings_json)
        }).execute()

    return {"success": True, "settings": data.settings_json}


# =====================================================
# Import / Export
# =====================================================

@router.get("/projects/{project_id}/episodes/template.csv")
async def get_import_template(
    project_id: str,
    authorization: str = Header(None)
):
    """Download CSV template for episode import."""
    user = await get_current_user_from_token(authorization)
    await verify_project_access(project_id, user["user_id"])

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "seasonNumber", "seasonTitle", "episodeNumber", "episodeCode",
        "title", "logline", "synopsis", "plannedRuntimeMinutes"
    ])
    # Example row
    writer.writerow(["1", "Season One", "1", "S01E01", "Pilot Episode", "A compelling logline", "Synopsis here", "45"])

    content = output.getvalue()
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=episode_import_template.csv"}
    )


@router.post("/projects/{project_id}/episodes/import")
async def import_episodes(
    project_id: str,
    file: UploadFile = File(...),
    authorization: str = Header(None)
):
    """Import episodes from CSV or XLSX file."""
    user = await get_current_user_from_token(authorization)
    project = await verify_project_access(project_id, user["user_id"])
    settings = await get_episode_settings(project_id)

    if not check_permission(project["user_role"], settings.get("canCreate", ["OWNER"])):
        raise HTTPException(status_code=403, detail="No permission to import episodes")

    filename = file.filename.lower()
    content = await file.read()

    rows = []

    if filename.endswith(".csv"):
        # Parse CSV
        text = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif filename.endswith(".xlsx"):
        # Parse XLSX
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                rows.append(row_dict)
        except ImportError:
            raise HTTPException(status_code=400, detail="XLSX support not available. Please use CSV format.")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use CSV or XLSX.")

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found")

    client = get_client()
    created = 0
    updated = 0
    errors = []

    # Cache seasons
    season_cache = {}

    for idx, row in enumerate(rows):
        try:
            season_num = int(row.get("seasonNumber") or 0)
            season_title = row.get("seasonTitle")
            episode_num = int(row.get("episodeNumber") or 1)
            episode_code = row.get("episodeCode")
            title = row.get("title")

            if not title:
                errors.append(f"Row {idx + 2}: Missing title")
                continue

            # Get or create season
            season_id = None
            if season_num > 0:
                cache_key = f"{season_num}"
                if cache_key in season_cache:
                    season_id = season_cache[cache_key]
                else:
                    season = execute_single(
                        "SELECT id FROM seasons WHERE project_id = :project_id AND season_number = :num",
                        {"project_id": project_id, "num": season_num}
                    )
                    if season:
                        season_id = str(season["id"])
                    else:
                        result = client.table("seasons").insert({
                            "project_id": project_id,
                            "season_number": season_num,
                            "title": season_title
                        }).execute()
                        if result.data:
                            season_id = result.data[0]["id"]
                    season_cache[cache_key] = season_id

            # Generate episode code if not provided
            if not episode_code:
                if season_id and season_num > 0:
                    episode_code = f"S{season_num:02d}E{episode_num:02d}"
                else:
                    episode_code = f"E{episode_num:02d}"

            # Check if episode exists (by code)
            existing = execute_single(
                "SELECT id FROM episodes WHERE project_id = :project_id AND episode_code = :code",
                {"project_id": project_id, "code": episode_code}
            )

            episode_data = {
                "project_id": project_id,
                "season_id": season_id,
                "episode_number": episode_num,
                "episode_code": episode_code,
                "title": title,
                "logline": row.get("logline"),
                "synopsis": row.get("synopsis"),
                "updated_at": datetime.utcnow().isoformat()
            }

            runtime = row.get("plannedRuntimeMinutes")
            if runtime:
                try:
                    episode_data["planned_runtime_minutes"] = int(runtime)
                except:
                    pass

            if existing:
                client.table("episodes").update(episode_data).eq("id", existing["id"]).execute()
                updated += 1
            else:
                episode_data["created_by_user_id"] = user["user_id"]
                client.table("episodes").insert(episode_data).execute()
                created += 1

        except Exception as e:
            errors.append(f"Row {idx + 2}: {str(e)}")

    return {
        "success": True,
        "created": created,
        "updated": updated,
        "errors": errors
    }


@router.get("/projects/{project_id}/episodes/export.csv")
async def export_episodes_csv(
    project_id: str,
    season_id: Optional[str] = None,
    authorization: str = Header(None)
):
    """Export episodes to CSV."""
    user = await get_current_user_from_token(authorization)
    await verify_project_access(project_id, user["user_id"])

    query = """
        SELECT e.*, s.season_number, s.title as season_title
        FROM episodes e
        LEFT JOIN seasons s ON s.id = e.season_id
        WHERE e.project_id = :project_id
    """
    params = {"project_id": project_id}

    if season_id:
        query += " AND e.season_id = :season_id"
        params["season_id"] = season_id

    query += " ORDER BY s.season_number NULLS LAST, e.episode_number"

    episodes = execute_query(query, params)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Season", "Episode #", "Code", "Title", "Pipeline Stage",
        "Edit Status", "Delivery Status", "Locked", "Logline",
        "Synopsis", "Planned Runtime", "Actual Runtime"
    ])

    for ep in episodes:
        writer.writerow([
            ep.get("season_number") or "",
            ep.get("episode_number"),
            ep.get("episode_code"),
            ep.get("title"),
            ep.get("pipeline_stage"),
            ep.get("edit_status"),
            ep.get("delivery_status"),
            "Yes" if ep.get("is_edit_locked") else "No",
            ep.get("logline") or "",
            ep.get("synopsis") or "",
            ep.get("planned_runtime_minutes") or "",
            ep.get("actual_runtime_minutes") or ""
        ])

    content = output.getvalue()
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=episodes_export.csv"}
    )


@router.get("/projects/{project_id}/episodes/print")
async def get_print_data(
    project_id: str,
    season_id: Optional[str] = None,
    authorization: str = Header(None)
):
    """Get data formatted for print view."""
    user = await get_current_user_from_token(authorization)
    await verify_project_access(project_id, user["user_id"])

    # Get project info
    project = execute_single(
        "SELECT title FROM backlot_projects WHERE id = :project_id",
        {"project_id": project_id}
    )

    # Get seasons
    season_query = "SELECT * FROM seasons WHERE project_id = :project_id ORDER BY season_number"
    params = {"project_id": project_id}
    if season_id:
        season_query = "SELECT * FROM seasons WHERE id = :season_id AND project_id = :project_id"
        params["season_id"] = season_id

    seasons = execute_query(season_query, params)

    # Get episodes grouped by season
    result_seasons = []
    for season in seasons:
        episodes = execute_query(
            """SELECT * FROM episodes
               WHERE season_id = :season_id
               ORDER BY episode_number""",
            {"season_id": season["id"]}
        )
        result_seasons.append({
            **dict(season),
            "episodes": [dict(e) for e in episodes]
        })

    # Get episodes without season
    if not season_id:
        no_season_episodes = execute_query(
            """SELECT * FROM episodes
               WHERE project_id = :project_id AND season_id IS NULL
               ORDER BY episode_number""",
            {"project_id": project_id}
        )
        if no_season_episodes:
            result_seasons.append({
                "id": None,
                "season_number": None,
                "title": "Standalone Episodes",
                "episodes": [dict(e) for e in no_season_episodes]
            })

    return {
        "project_title": project["title"] if project else "Unknown Project",
        "generated_at": datetime.utcnow().isoformat(),
        "seasons": result_seasons
    }
